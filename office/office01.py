from openpyxl import load_workbook

wb = load_workbook("file/third/template.xlsx")  # 打开一个xlsx文件。
print(wb.sheetnames)
sheet = wb["Sheet3"]  # 看看打开的excel表里面有哪些sheet页。
# 下面读取到指定的Sheet页
print(sheet["C"])
print(sheet["4"])
print(sheet["C4"].value)  # c4     <-第C4格的值
print(sheet.max_row)  # 10     <-最大行数
print(sheet.max_column)  # 5     <-最大列数
for i in sheet["C"]:
    print(i.value, end=" ")  # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值
