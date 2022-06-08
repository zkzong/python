# author: zong
# date: 2022/6/4 9:06

import openpyxl
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, GradientFill, Side, Border, Alignment, NamedStyle
from openpyxl.utils import FORMULAE

wb = openpyxl.Workbook()
ws = wb.active
print(ws.title)
# 第一行第一列
ws['A1'] = 520
# 第二行
ws.append([1, 2, 3])
ws['A3'] = datetime.datetime.now()
wb.save('demo.xlsx')


# 打开excel文件
wb = openpyxl.load_workbook('豆瓣TOP250电影.xlsx')
print(type(wb))

# 获取工作表
# print(wb.get_sheet_names())
print(wb.sheetnames)
# ws = wb.get_sheet_by_name('Sheet')
ws = wb['Sheet']

# 创建和删除工作表
nws = wb.create_sheet(index=0, title='新建工作表')
print(wb.sheetnames)
wb.remove(wb['新建工作表'])
print(wb.sheetnames)

# 定位单元格
c = ws['A2']
print(c.row)
print(c.column)
print(c.coordinate)
print(ws['A2'].value)
print(c.value)
d = c.offset(2, 0)
print(d.value)

# 'AAA'是多少
print(openpyxl.cell.cell.get_column_letter(496))
print(openpyxl.utils.cell.column_index_from_string('JB'))

# 访问多个单元格
for each_movie in ws['A2':'B10']:
    for each_cell in each_movie:
        print(each_cell.value, end=' ')
    print('\n')
for each_row in ws.rows:
    print(each_row[0].value)
for each_row in ws.iter_rows(min_row=2, min_col=1, max_row=4, max_col=2):
    print(each_row[0].value)

# 拷贝工作表
new = wb.copy_worksheet(ws)
print(type(new))
wb.save('豆瓣TOP250电影.xlsx')

# 个性化工作表标签栏
wb = openpyxl.Workbook()
ws1 = wb.create_sheet(title='工作表1')
ws2 = wb.create_sheet(title='工作表2')
ws3 = wb.create_sheet(title='工作表3')
ws4 = wb.create_sheet(title='工作表4')
ws1.sheet_properties.tabColor = 'FF0000'
ws2.sheet_properties.tabColor = '00FF00'
ws3.sheet_properties.tabColor = '0000FF'
ws4.sheet_properties.tabColor = '8B008B'
wb.save('demo.xlsx')

# 调整行高和列宽
ws2.row_dimensions[2].height = 100
ws2.column_dimensions['c'].width = 50
wb.save('demo.xlsx')

# 合并和拆分单元格
ws1.merge_cells('A1:C3')
ws1['A1'] = '这是一个合并单元格'
wb.save('demo.xlsx')
ws1.unmerge_cells('A1:C3')
wb.save('demo.xlsx')

# 冻结窗口
openpyxl.load_workbook('demo.xlsx')
ws = wb.active
# B8单元格上面和左面的被冻结
ws.freeze_panes = 'B8'
wb.save('demo.xlsx')
# 解冻
ws.freeze_panes = 'A1'
ws.freeze_panes = None
wb.save('demo.xlsx')

# 设置单元格字体
wb = Workbook()
ws = wb.active
b2 = ws['B2']
b2.value = 'Hello'
bold_red_font = Font(bold=True, color='FF0000')
b2.font = bold_red_font
b3 = ws['B3']
b3.value = 'World'
italic_strike_blue_16font = Font(size=16, italic=True, strike=True, color='0000FF')
b3.font = italic_strike_blue_16font
wb.save('demo1.xlsx')

# 填充单元格
yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')
b2.fill = yellow_fill
wb.save('demo2.xlsx')

red2green_fill = GradientFill(type='linear', stop=('FF0000', '00FF00'))
b3.fill = red2green_fill
wb.save('demo2.xlsx')

# 设置边框
thin_side = Side(border_style='thin', color='000000')
double_side = Side(border_style='double', color='FF0000')
b2.border = Border(diagonal=thin_side, diagonalUp=True, diagonalDown=True)
b3.border = Border(top=double_side, left=double_side, right=double_side, bottom=double_side)
wb.save('demo3.xlsx')

# 文本对齐
ws.merge_cells('A1:C2')
ws['A1'] = '这是一个合并单元格'
center_alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].alignment = center_alignment
wb.save('demo4.xlsx')

# 命名样式
'''
使用命名样式只需要四个步骤：
1. 实例化一个NameStyle类
2. 初始化命名样式
3. 注册命名样式到工作簿中
4. 将单元格的style属性赋值为命名样式
'''
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True, size=20)
highlight.alignment = Alignment(horizontal='center', vertical='center')
wb.add_named_style(highlight)
ws['A1'].style = highlight
ws['B5'].value = 'Hello'
ws['B5'].style = highlight
wb.save('demo5.xlsx')

# 数字格式
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['文本', '数字'])
ws['A2'] = '520'
ws['B2'] = 520
wb.save('test1.xlsx')

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 88.8
ws['A1'].number_format = '#,###.00元'
ws['A2'] = datetime.datetime.today()
ws['A2'].number_format = 'yyyy-mm-dd'
wb.save('test2.xlsx')

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'].number_format = '[RED]+#,###.00;[GREEN]-#,###.00'
ws['A1'] = 99
ws['A2'].number_format = '[RED]+#,###.00;[GREEN]-#,###.00'
ws['A2'] = -99
ws['A3'].number_format = '[RED];[GREEN];[BLUE];[YELLOW]'
ws['A3'] = 0
ws['A4'].number_format = '[RED];[GREEN];[BLUE];[YELLOW]'
ws['A4'] = 'Hello'
ws['A5'].number_format = '[=1]男;[=0]女'
ws['A5'] = 0
ws['A6'].number_format = '[=1]男;[=0]女'
ws['A6'] = 1
ws['A7'].number_format = '[=1]男;[=0]女'
ws['A7'] = 2
ws['A8'].number_format = '[<60][RED]不及格;[>=60][GREEN]及格'
ws['A8'] = 58
ws['A9'].number_format = '[<60][RED]不及格;[>=60][GREEN]及格'
ws['A9'] = 68
wb.save('number.xlsx')

# 函数公式
print('SUM' in FORMULAE)
print('SAM' in FORMULAE)