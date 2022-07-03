import openpyxl
import time

ls = [['马坡','接入交换','192.168.1.1','G0/3','AAAA-AAAA-AAAA'],
      ['马坡','接入交换','192.168.1.2','G0/8','BBBB-BBBB-BBBB'],
      ['马坡','接入交换','192.168.1.2','G0/8','CCCC-CCCC-CCCC'],
      ['马坡','接入交换','192.168.1.2','G0/8','DDDD-DDDD-DDDD']]

##定义数据

time_format = '%Y-%m-%d__%H:%M:%S'
time_current = time.strftime(time_format)
##定义时间格式

def savetoexcel(data,sheetname,wbname):
    print("写入excel：")
    wb=openpyxl.load_workbook(filename=wbname)
    ##打开excel文件

    sheet=wb.active #关联excel活动的Sheet（这里关联的是Sheet1）
    max_row = sheet.max_row #获取Sheet1中当前数据最大的行数
    row = max_row + 3   #将新数据写入最大行数+3的位置
    data_len=row+len(data)  #计算当前数据长度

    for data_row in range(row,data_len):  # 写入数据
    ##轮询每一行进行写入数据。
        for data_col1 in range(2,7):
        ##针对每一行下面还要进行for循环来写入列的数据
            _ =sheet.cell(row=data_row, column=1, value=str(time_current))
            ##每行第一列写入时间
            _ =sheet.cell(row=data_row,column=data_col1,value=str(data[data_row-data_len][data_col1-2]))
            #从第二列开始写入数据

    wb.save(filename=wbname)    #保存数据
    print("保存成功")

savetoexcel(ls,"Sheet1","template.xlsx")
