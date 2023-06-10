import openpyxl
import pandas as pd

# 1、pandas
# 获取所有sheet名称
df = pd.read_excel('file/employee.xlsx', sheet_name=None)
print(list(df))
for i in df.keys():
    print(i)

# 默认读取第一个sheet
df = pd.read_excel('file/employee.xlsx', header=None)
# 行数
rows = len(df)
print(rows)
# 列数
columns = len(df.columns)
print(columns)
# 行数和列数
print(df.shape)

for row in range(rows):
    for column in range(columns):
        print(df.iloc[row, column], end=' ')
    print()
for row in df.index.values:
    print(df.iloc[row, 0], end=' ')
    print(df.iloc[row, 1], end=' ')
    print(df.iloc[row, 2], end=' ')
    print(df.iloc[row, 3], end=' ')
    print(df.iloc[row, 4], end=' ')
    print()

# 读取指定sheet
# sheet = pd.read_excel('file/employee.xlsx', 'Sheet2')
# print(type(sheet.values))
# print(sheet.values)
# print(sheet.index)
# print(sheet.info)
# print(sheet.iloc)
# print(sheet.loc)
# print(sheet.index.values)

# 2、openpyxl
wb = openpyxl.load_workbook('file/employee.xlsx')
sheet = wb.active
print(sheet)
# 获取所有sheet
sheetnames = wb.sheetnames
for sheetname in sheetnames:
    print(sheetname)
# 根据name获取sheet
# deprecated
# sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb['Sheet1']
print(sheet)
# sheet = wb.get_sheet_by_name('Sheet2')
# sheet = wb['Sheet2']
# print(sheet)
# 最小行
minrow = sheet.min_row
print('min_row=' + str(minrow))
# 最大行
maxrow = sheet.max_row
print('max_row=' + str(maxrow))
# 最小列
mincol = sheet.min_column
print('min_column=' + str(mincol))
# 最大列
maxcol = sheet.max_column
print('max_column=' + str(maxcol))
# 按行读取
for i in range(minrow, maxrow + 1):
    for j in range(mincol, maxcol + 1):
        cell = sheet.cell(i, j).value
        print(cell, end=" ")
    print()
# 按列读取
for m in range(mincol, maxcol + 1):
    for n in range(minrow, maxrow + 1):
        cell = sheet.cell(n, m).value
        print(cell, end=" ")
    print()
